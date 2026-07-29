from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
from typing import List, Dict, Any, Optional
from utils.datetime_utils import get_pakistan_time

class ExcelReportService:
    """Service for generating Excel reports for sales data"""
    
    @staticmethod
    def generate_sales_report(
        sales: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Generate an Excel report for sales data with applied filters
        
        Args:
            sales: List of sale documents
            filters: Dictionary containing filter information (doctor_name, date_from, date_to, etc.)
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report title
        ws.merge_cells('A1:M1')
        ws['A1'] = "Sales Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Add filter information
        current_row = 2
        if filters:
            ws.merge_cells(f'A{current_row}:M{current_row}')
            filter_text = "Filters Applied: "
            filter_parts = []
            
            if filters.get('doctor_name'):
                filter_parts.append(f"Doctor: {filters['doctor_name']}")
            if filters.get('date_from'):
                filter_parts.append(f"From: {filters['date_from']}")
            if filters.get('date_to'):
                filter_parts.append(f"To: {filters['date_to']}")
            
            if filter_parts:
                filter_text += ", ".join(filter_parts)
            else:
                filter_text += "None"
            
            ws[f'A{current_row}'] = filter_text
            ws[f'A{current_row}'].font = Font(italic=True, size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="left")
            current_row += 1
        
        # Add generation date
        ws.merge_cells(f'A{current_row}:M{current_row}')
        pakistan_now = get_pakistan_time()
        ws[f'A{current_row}'] = f"Generated on: {pakistan_now.strftime('%Y-%m-%d %I:%M:%S %p')}"
        ws[f'A{current_row}'].font = Font(italic=True, size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal="right")
        current_row += 2
        
        # Define headers
        headers = [
            "Sale ID",
            "Date",
            "Time",
            "Patient Name",
            "Phone",
            "Referred By",
            "Test(s)",
            "Doctor",
            "Subtotal",
            "Cost Price",
            "Discount",
            "Final Amount",
            "Doctor Share",
            "Total Profit",
            "Net Profit (After Doctor Share)",
            "Payment Status"
        ]
        
        # Write headers
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        # Write data
        total_subtotal = 0
        total_cost = 0
        total_discount = 0
        total_final = 0
        total_commission = 0
        total_profit = 0
        total_net_profit = 0
        
        for sale in sales:
            # Extract date and time (already in Pakistan timezone from database)
            created_at = sale.get('created_at')
            if isinstance(created_at, datetime.datetime):
                date_str = created_at.strftime('%Y-%m-%d')
                time_str = created_at.strftime('%I:%M:%S %p')
            else:
                date_str = 'N/A'
                time_str = 'N/A'
            
            # Get test names
            test_name = sale.get('test_name', 'N/A')
            
            # Get referred by (hospital_name)
            referred_by = sale.get('hospital_name', 'N/A')
            
            # Calculate amounts
            subtotal = float(sale.get('subtotal', 0))
            discount = float(sale.get('discount_amount', 0))
            final_amount = float(sale.get('final_amount', 0))
            commission = float(sale.get('doctor_commission_amount', 0))
            
            # Calculate total cost from tests or from stored cost_price
            total_cost_value = float(sale.get('cost_price', 0))
            if total_cost_value == 0:
                # Fallback: calculate from tests if not stored
                tests = sale.get('tests', [])
                for test in tests:
                    cost = float(test.get('cost_price', 0))
                    total_cost_value += cost
            
            # Calculate profits according to dashboard logic:
            # Total Profit = Final Amount - Cost Price
            profit = final_amount - total_cost_value
            
            # Net Profit (After Commission) = Total Profit - Doctor Commission
            net_profit = profit - commission
            
            # Payment status
            is_partial = sale.get('is_partial', False)
            paid_amount = float(sale.get('paid_amount', 0))
            remaining_amount = float(sale.get('remaining_amount', 0))
            
            if is_partial and remaining_amount > 0:
                payment_status = f"Partial (Paid: Rs {paid_amount:.2f}, Remaining: Rs {remaining_amount:.2f})"
            else:
                payment_status = "Fully Paid"
            
            # Write row data with all columns including Referred By and Cost Price
            row_data = [
                sale.get('sale_id', 'N/A'),
                date_str,
                time_str,
                sale.get('patient_name', 'N/A'),
                sale.get('phone', 'N/A'),
                referred_by,
                test_name,
                sale.get('doctor_name', 'N/A'),
                subtotal,
                total_cost_value,
                discount,
                final_amount,
                commission,
                profit,
                net_profit,
                payment_status
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                
                # Format currency columns
                if col_num in [9, 10, 11, 12, 13, 14, 15]:  # Subtotal, Cost, Discount, Final Amount, Commission, Total Profit, Net Profit
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Add to totals
            total_subtotal += subtotal
            total_cost += total_cost_value
            total_discount += discount
            total_final += final_amount
            total_commission += commission
            total_profit += profit
            total_net_profit += net_profit
            
            current_row += 1
        
        # Add totals row
        current_row += 1
        ws.merge_cells(f'A{current_row}:H{current_row}')
        total_cell = ws.cell(row=current_row, column=1)
        total_cell.value = "TOTAL"
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = border
        
        # Total amounts: Subtotal, Cost Price, Discount, Final Amount, Commission, Total Profit, Net Profit
        for col_num, total_value in enumerate([total_subtotal, total_cost, total_discount, total_final, total_commission, total_profit, total_net_profit], 9):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = total_value
            cell.font = Font(bold=True, size=12)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Payment Status column in totals row
        cell = ws.cell(row=current_row, column=16)
        cell.value = f"{len(sales)} sales"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # Adjust column widths
        column_widths = {
            'A': 10,  # Sale ID
            'B': 12,  # Date
            'C': 10,  # Time
            'D': 20,  # Patient Name
            'E': 15,  # Phone
            'F': 20,  # Referred By
            'G': 35,  # Test(s)
            'H': 20,  # Doctor
            'I': 12,  # Subtotal
            'J': 12,  # Cost Price
            'K': 12,  # Discount
            'L': 12,  # Final Amount
            'M': 12,  # Commission
            'N': 12,  # Total Profit
            'O': 20,  # Net Profit (After Commission)
            'P': 30,  # Payment Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = f'A{header_row + 1}'
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
