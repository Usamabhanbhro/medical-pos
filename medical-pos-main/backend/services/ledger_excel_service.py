from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
from typing import List, Dict, Any, Optional
import sys
import os

# Add parent directory to path to import utils
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.datetime_utils import get_pakistan_time


class LedgerExcelReportService:
    """Service for generating Excel reports for ledger data"""
    
    @staticmethod
    def generate_ledger_report(
        entries: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Generate an Excel report for ledger entries with applied filters
        
        Args:
            entries: List of ledger entry documents
            filters: Dictionary containing filter information (category, date_from, date_to)
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger Report"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Ledger Expense Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Add filter information
        current_row = 2
        if filters:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            filter_text = "Filters Applied: "
            filter_parts = []
            
            if filters.get('category'):
                filter_parts.append(f"Category: {filters['category']}")
            if filters.get('date_from'):
                filter_parts.append(f"From: {filters['date_from']}")
            if filters.get('date_to'):
                filter_parts.append(f"To: {filters['date_to']}")
            
            if filter_parts:
                filter_text += ", ".join(filter_parts)
            else:
                filter_text += "None"
            
            ws[f'A{current_row}'] = filter_text
            ws[f'A{current_row}'].font = Font(italic=True, size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="left")
            current_row += 1
        
        # Add generation date
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"Generated on: {get_pakistan_time().strftime('%Y-%m-%d %I:%M:%S %p')}"
        ws[f'A{current_row}'].font = Font(italic=True, size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal="right")
        current_row += 2
        
        # Define headers
        headers = [
            "Date",
            "Category",
            "Description",
            "Amount",
            "Notes",
            "Created At"
        ]
        
        # Write headers
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        # Write data
        total_amount = 0
        official_total = 0
        unofficial_total = 0
        
        for entry in entries:
            # Extract date (already in Pakistan timezone from database)
            entry_date = entry.get('date')
            if isinstance(entry_date, datetime.datetime):
                date_str = entry_date.strftime('%Y-%m-%d')
            else:
                date_str = 'N/A'
            
            # Extract created_at (already in Pakistan timezone from database)
            created_at = entry.get('created_at')
            if isinstance(created_at, datetime.datetime):
                created_str = created_at.strftime('%Y-%m-%d %I:%M:%S %p')
            else:
                created_str = 'N/A'
            
            # Get data
            category = entry.get('category', 'N/A')
            description = entry.get('description', 'N/A')
            amount = float(entry.get('amount', 0))
            notes = entry.get('notes', '')
            
            # Write row data
            row_data = [
                date_str,
                category.capitalize() if category else 'N/A',
                description,
                amount,
                notes or '',
                created_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                
                # Format amount column
                if col_num == 4:  # Amount
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num == 2:  # Category
                    cell.alignment = Alignment(horizontal="center")
                    # Color code based on category
                    if category == 'official':
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Add to totals
            total_amount += amount
            if category == 'official':
                official_total += amount
            else:
                unofficial_total += amount
            
            current_row += 1
        
        # Add summary section
        current_row += 2
        
        # Official Total
        ws.merge_cells(f'A{current_row}:C{current_row}')
        summary_cell = ws.cell(row=current_row, column=1)
        summary_cell.value = "Official Expenses Total"
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.alignment = Alignment(horizontal="right", vertical="center")
        summary_cell.border = border
        summary_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        cell = ws.cell(row=current_row, column=4)
        cell.value = official_total
        cell.font = Font(bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        current_row += 1
        
        # Unofficial Total
        ws.merge_cells(f'A{current_row}:C{current_row}')
        summary_cell = ws.cell(row=current_row, column=1)
        summary_cell.value = "Unofficial Expenses Total"
        summary_cell.font = Font(bold=True, size=11)
        summary_cell.alignment = Alignment(horizontal="right", vertical="center")
        summary_cell.border = border
        summary_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        cell = ws.cell(row=current_row, column=4)
        cell.value = unofficial_total
        cell.font = Font(bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        current_row += 1
        
        # Grand Total
        ws.merge_cells(f'A{current_row}:C{current_row}')
        total_cell = ws.cell(row=current_row, column=1)
        total_cell.value = "GRAND TOTAL"
        total_cell.font = Font(bold=True, size=12, color="FFFFFF")
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = border
        total_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        cell = ws.cell(row=current_row, column=4)
        cell.value = total_amount
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Entry count
        ws.merge_cells(f'E{current_row}:F{current_row}')
        cell = ws.cell(row=current_row, column=5)
        cell.value = f"{len(entries)} entries"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Adjust column widths
        column_widths = {
            'A': 12,  # Date
            'B': 15,  # Category
            'C': 40,  # Description
            'D': 15,  # Amount
            'E': 40,  # Notes
            'F': 20,  # Created At
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = f'A{header_row + 1}'
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
